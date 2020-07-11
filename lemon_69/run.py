#执行文件
#获取到所有的测试数据
#将R_W_EXCEL改成R_W_excel
# from lemon_69.R_W_excel import read_data
# from lemon_69.http_request import http_request
# from openpyxl import load_workbook
# all_case=read_data('test_case.xlsx','recharge')
# print('获取到的所有测试数据是:',all_case)

#在R_W_excel.py里面:加上if_name_=='_main_':      #输入main
#当鼠标在当前文件下，才会执行当前代码
#写个main函数
# all_case=read_data('test_case.xlsx','recharge')
# print('获取到的所有测试数据是:',all_case)

#执行测试（文件名与函数名可以一致）--先执行第一条
# print('第一条用例的数据：',all_case[0])
# # test_data=all_case[0]
# # method=test_data[3]
# # uri=test_data[4]
# # param=eval(test_data[5])
# # # #从excel表里读出来的数据是字符串类型的  发送请求希望param是字典dict类型
# # expected=eval(test_data[6])
# # #print(type(param))  可以去掉
# # ip='http://120.78.128.25:8766'
# # url=ip+uri
# # header={'X-Lemonban-Media-Type':'lemonbanban.v2'}  #去掉
# # response=http_request(url,param,header,method)
# # #                                   可以省掉
# # print('最后的结果值:',response)
# # # #由于登陆与充值的请求头不一样 所以需要优化一下 token值可以参数化 请求数据不需要参数化

#登陆充值一起跑，区分不出来哪个是充值需要的token值，全部做普通的处理
# print('第一条用例的数据:',all_case[0])
# for i in range(len(all_case)):
#     test_data = all_case[i]
#     expected = eval(test_data[6])
#     ip = 'http://120.78.128.25:8766'
#     response = http_request(ip + test_data[4], eval(test_data[5], token='Bearer ' + token, method=test_data[3]))
# print('最后的结果值:', response)


#执行测试---先执行第一条
# print('第一条用例的数据:',all_case[0])
# login_case_data=all_case[0]
# ip='http://120.78.128.25:8766'
# log_response=http_request(ip+login_case_data[4],eval(login_case_data[5]),token=None,method=login_case_data[3])
# print('最后的结果值:',log_response)
#
#
# for i in range(1,len(all_case)):
#     test_data=all_case[i]
#     token=log_response['data']['token_info']['token']
#     response=http_request(ip+test_data[4],eval(test_data[5]),token='Bearer '+token,method=test_data[3])
#     expected=eval(test_data[6])   #期望值
#     print('最后的结果值:', response)
#



#1.完成充值的测试用例设计
#2.把http请求全部复用为一个函数
#3.写一个函数专门用来读取测试用例数据，对数据进行封装成嵌套列表
#4.写一个run函数来执行测试用例

#1.保证每个用例之间的数据是独立的，每行为一个用例，每一个列表为一个用例，把它放在一个单独的列表里面，每个列表代表一条测试用例
#2.一次性读取出来更加方便

#全局变量  局部变量
#函数内的变量---局部的    函数外---全局的

# Token=None    #全局变量，初始值设置为None
# def run():
#     global Token   #在这里声明   函数外的Token和函数内的Token是同一个值
#     all_case=read_data('test_case.xlsx','recharge')
#     print('获取到的所有测试数据是:',all_case)
#     for i in range(len(all_case)):       #在http_request进行请求的时候，判断是否是登陆请求
#         test_data=all_case[i]
#         #if test_data[0]=1
#         #if test_data[1]='登陆'             #判断两边是否相等 比较运算符
#         ip='http://120.78.128.25:8766'
#         response=http_request(ip+test_data[4],eval(test_data[5]),token=Token,method=test_data[3])
#         if 'login' in test_data[4]:         #成员运算符  #它就是一个登陆的用例
#             Token='Bearer '+response['data']['token_info']['token']
#         print('最后的结果值:',response)
#         # return response.json()




#  #开始写入结果
# wb=load_workbook('test_case.xlsx')
# sheet=wb['recharge']
#     #定位单元格存值  行 列 值
# sheet.cell(row=test_data[0]+1,column=8).value=str(response)
# #进行判断，期望值与实际值是否相等，判断这个用例是否执行通过
# # actual={'code':response['code'],'msg':response['msg']}
# # if eval(test_data[6])==actual:
# #      print('测试用例执行通过')
# #      sheet.cell(row=test_data[0]+1,column=9).value='PASS'
# # else:
# #     print('测试用例执行不通过')
# #     sheet.cell(row=test_data[0] + 1, column=9).value ='FAIL'
# #保存
# # wb.save('test_case.xlsx')


# #调用函数
# run()
#
#
# Token=None    #全局变量，初始值设置为None
# def run(file_name,sheet_name,c1,c2):
#     global Token   #在这里声明   函数外的Token和函数内的Token是同一个值
#     all_case=read_data(file_name,sheet_name)
#     print('获取到的所有测试数据是:',all_case)
#     for test_data in all_case:   #在http_request进行请求的时候，判断是否是登陆请求
#         ip = 'http://120.78.128.25:8766'
#         response = http_request(ip + test_data[4], eval(test_data[5]), token=Token, method=test_data[3])
#         if 'login' in test_data[4]:         #成员运算符  #它就是一个登陆的用例
#             Token='Bearer '+response['data']['token_info']['token']
#         print('最后的结果值:',response)
#
# #开始写入结果
# write_data(file_name,sheet_name,test_data[0]+1,c1,str(response))
# #进行判断，期望值与实际值是否相等，判断这个用例是否执行通过
# # actual={'code':response['code'],'msg':response['msg']}
# # if eval(test_data[6])==actual:
# #      print('测试用例执行通过')
# #      write_data(file_name,sheet_name,test_data[0]+1,c2,'PASS')
# # else:
# #     print('测试用例执行不通过')
# #      write_data(file_name,sheet_name,test_data[0]+1,c2,'FAIL')
# # wb.save('test_case.xlsx')
#
#
# #调用函数
# #执行的充值的接口
# run('test_case.xlsx','recharge',8,9)
# #执行提现的接口
# run('test_case.xlsx','withdraw',8,9)





from lemon_69.R_W_excel import read_data
from lemon_69.http_request import http_request
from openpyxl import load_workbook

Token=None    #全局变量，初始值设置为None
def run():
    global Token   #在这里声明   函数外的Token和函数内的Token是同一个值
    all_case=read_data('test_case.xlsx','recharge')
    print('获取到的所有测试数据是:',all_case)
    for i in range(len(all_case)):       #在http_request进行请求的时候，判断是否是登陆请求
        test_data=all_case[i]
        #if test_data[0]=1
        #if test_data[1]='登陆'             #判断两边是否相等 比较运算符
        ip='http://120.78.128.25:8766'
        response=http_request(ip+test_data[4],eval(test_data[5]),token=Token,method=test_data[3])
        if 'login' in test_data[4]:         #成员运算符  #它就是一个登陆的用例
            Token='Bearer '+response['data']['token_info']['token']
        print('最后的结果值:', response)
        wb=load_workbook('test_case.xlsx')
        sheet=wb['recharge']
        sheet.cell(row=test_data[0]+1,column=8).value=str(response)
        wb.save('test_case.xlsx')

#调用函数
run()