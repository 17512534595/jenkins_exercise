#用来读取测试数据
from openpyxl import load_workbook
# wb=load_workbook('test_case.xlsx')
# sheet=wb['recharge']
# all_case=[]   #存储所有行的测试用例数据
# for i in range(2,sheet.max_row+1):
#     case=[]    #某一行测试用例数据
#     for j in range(1,sheet.max_column-1):
#         case.append(sheet.cell(row=i,column=j).value)
#         print(sheet.cell(row=i,column=j).value)
#     print(case)
#     all_case.append(case)
# print(all_case)





def read_data(file_name,sheet_name):  #读取数据的函数
    wb = load_workbook(file_name)
    sheet = wb[sheet_name]
    all_case=[]
    for i in range(2, sheet.max_row+1):
        case = []
        for j in range(1, sheet.max_column - 1):
            case.append(sheet.cell(row=i, column=j).value)
        all_case.append(case)
    return all_case   #返回所有测试用例数据

# def write_data(file_name,sheet_name,row,column,value):   #此函数是写入结果到excel中
#     wb=load_workbook(file_name)
#     sheet=wb[sheet_name]
#     #定位单元格存值 行 列 值
#     sheet.cell(row=row,column=column).value=value

#保存
# wb.save(file_name)


if __name__ == '__main__':
    all_case=read_data('test_case.xlsx','recharge')
    print('所有的测试数据为:',all_case)


